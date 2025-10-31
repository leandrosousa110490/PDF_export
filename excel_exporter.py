"""
Excel Export Utility for PDF Value Extraction System
====================================================

This module provides functionality to export extraction results to Excel files
with proper formatting and multiple output options.
"""

import os
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from config import (
    EXCEL_COLUMNS, EXCEL_AUTO_ADJUST_COLUMNS, EXCEL_FREEZE_HEADER, 
    EXCEL_ADD_FILTERS, get_full_excel_path, get_excel_output_path
)

class ExcelExporter:
    """Class to handle Excel export functionality"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.results = []
    
    def add_result(self, pdf_name: str, config_name: str, extracted_value: str, 
                   status: str = "Success", timestamp: datetime = None):
        """Add a single extraction result"""
        if timestamp is None:
            timestamp = datetime.now()
        
        result = {
            'pdf_name': pdf_name,
            'config_name': config_name,
            'extracted_value': extracted_value,
            'extraction_status': status,
            'timestamp': timestamp.strftime("%Y-%m-%d %H:%M:%S")
        }
        self.results.append(result)
    
    def add_results_batch(self, results: List[Dict[str, Any]]):
        """Add multiple extraction results at once"""
        for result in results:
            self.add_result(
                pdf_name=result.get('pdf_name', ''),
                config_name=result.get('config_name', ''),
                extracted_value=result.get('extracted_value', ''),
                status=result.get('status', 'Success'),
                timestamp=result.get('timestamp')
            )
    
    def create_workbook(self):
        """Create a new Excel workbook with formatted headers"""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Extraction Results"
        
        # Add headers
        headers = [
            EXCEL_COLUMNS['pdf_name'],
            EXCEL_COLUMNS['config_name'],
            EXCEL_COLUMNS['extracted_value'],
            EXCEL_COLUMNS['extraction_status'],
            EXCEL_COLUMNS['timestamp']
        ]
        
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            # Format header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def populate_data(self):
        """Populate the worksheet with extraction results"""
        if not self.results:
            return
        
        for row, result in enumerate(self.results, 2):  # Start from row 2 (after header)
            self.worksheet.cell(row=row, column=1, value=result['pdf_name'])
            self.worksheet.cell(row=row, column=2, value=result['config_name'])
            self.worksheet.cell(row=row, column=3, value=result['extracted_value'])
            self.worksheet.cell(row=row, column=4, value=result['extraction_status'])
            self.worksheet.cell(row=row, column=5, value=result['timestamp'])
            
            # Color-code status column
            status_cell = self.worksheet.cell(row=row, column=4)
            if result['extraction_status'].lower() == 'success':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif result['extraction_status'].lower() in ['failed', 'error', 'no match']:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    def format_worksheet(self):
        """Apply formatting to the worksheet"""
        if not self.worksheet:
            return
        
        # Auto-adjust column widths
        if EXCEL_AUTO_ADJUST_COLUMNS:
            for column in self.worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                self.worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        if EXCEL_FREEZE_HEADER:
            self.worksheet.freeze_panes = "A2"
        
        # Add filters
        if EXCEL_ADD_FILTERS:
            self.worksheet.auto_filter.ref = self.worksheet.dimensions
    
    def add_summary_sheet(self):
        """Add a summary sheet with statistics"""
        summary_sheet = self.workbook.create_sheet("Summary")
        
        # Calculate statistics
        total_extractions = len(self.results)
        successful_extractions = len([r for r in self.results if r['extraction_status'].lower() == 'success'])
        failed_extractions = total_extractions - successful_extractions
        success_rate = (successful_extractions / total_extractions * 100) if total_extractions > 0 else 0
        
        # Unique PDFs and configs
        unique_pdfs = len(set(r['pdf_name'] for r in self.results))
        unique_configs = len(set(r['config_name'] for r in self.results))
        
        # Add summary data
        summary_data = [
            ["Extraction Summary", ""],
            ["", ""],
            ["Total Extractions", total_extractions],
            ["Successful Extractions", successful_extractions],
            ["Failed Extractions", failed_extractions],
            ["Success Rate", f"{success_rate:.1f}%"],
            ["", ""],
            ["Unique PDF Files", unique_pdfs],
            ["Unique Configurations", unique_configs],
            ["", ""],
            ["Generated On", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            summary_sheet.cell(row=row, column=1, value=label)
            summary_sheet.cell(row=row, column=2, value=value)
            
            # Format header
            if row == 1:
                cell = summary_sheet.cell(row=row, column=1)
                cell.font = Font(bold=True, size=14)
        
        # Auto-adjust columns
        for column in summary_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = max_length + 2
            summary_sheet.column_dimensions[column_letter].width = adjusted_width
    
    def save_to_file(self, filename: str = None) -> str:
        """Save the workbook to an Excel file"""
        if filename is None:
            filename = get_full_excel_path()
        
        # Ensure output directory exists (only if filename has a directory path)
        dir_path = os.path.dirname(filename)
        if dir_path:  # Only create directory if there's a directory path
            os.makedirs(dir_path, exist_ok=True)
        
        # Create and format workbook
        self.create_workbook()
        self.populate_data()
        self.format_worksheet()
        self.add_summary_sheet()
        
        # Save file
        self.workbook.save(filename)
        return filename
    
    def export_results(self, results: List[Dict[str, Any]] = None, filename: str = None) -> str:
        """Complete export process - add results and save to file"""
        if results:
            self.add_results_batch(results)
        
        return self.save_to_file(filename)
    
    def clear_results(self):
        """Clear all stored results"""
        self.results = []
    
    def get_results_count(self) -> int:
        """Get the number of stored results"""
        return len(self.results)
    
    def print_summary(self):
        """Print a summary of the results"""
        if not self.results:
            print("No results to summarize.")
            return
        
        total = len(self.results)
        successful = len([r for r in self.results if r['extraction_status'].lower() == 'success'])
        failed = total - successful
        
        print(f"\nExtraction Results Summary:")
        print(f"Total extractions: {total}")
        print(f"Successful: {successful}")
        print(f"Failed: {failed}")
        print(f"Success rate: {(successful/total*100):.1f}%")
        
        unique_pdfs = len(set(r['pdf_name'] for r in self.results))
        unique_configs = len(set(r['config_name'] for r in self.results))
        print(f"Unique PDF files: {unique_pdfs}")
        print(f"Unique configurations: {unique_configs}")

def quick_export(results: List[Dict[str, Any]], filename: str = None) -> str:
    """Quick function to export results to Excel"""
    exporter = ExcelExporter()
    return exporter.export_results(results, filename)

# Example usage
if __name__ == "__main__":
    # Example results
    sample_results = [
        {
            'pdf_name': 'sample_invoice_1.pdf',
            'config_name': 'Invoice Total',
            'extracted_value': '$1,234.56',
            'status': 'Success'
        },
        {
            'pdf_name': 'sample_invoice_2.pdf',
            'config_name': 'Invoice Total',
            'extracted_value': 'No match found',
            'status': 'No Match'
        },
        {
            'pdf_name': 'sample_invoice_1.pdf',
            'config_name': 'Customer Name',
            'extracted_value': 'John Doe',
            'status': 'Success'
        }
    ]
    
    # Export to Excel
    exporter = ExcelExporter()
    filename = exporter.export_results(sample_results)
    print(f"Results exported to: {filename}")
    exporter.print_summary()